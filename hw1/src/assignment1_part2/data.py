from datetime import datetime, timedelta
from io import BytesIO
from zipfile import ZipFile

from openpyxl.reader.excel import load_workbook


def to_float(value):
    if value in (None, ""):
        return 0.0
    return float(value)


def excel_to_datetime(value):
    if isinstance(value, datetime):
        return value
    if isinstance(value, (int, float)):
        return datetime(1899, 12, 30) + timedelta(days=float(value))
    if isinstance(value, str):
        for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d"):
            try:
                return datetime.strptime(value, fmt)
            except ValueError:
                pass
    raise ValueError(f"unsupported datetime value: {value!r}")


def log_price_series(dates, prices):
    clean_dates = []
    clean_prices = []
    for dt, price in zip(dates, prices):
        price = to_float(price)
        if price > 0:
            clean_dates.append(dt)
            clean_prices.append(price)
    return clean_dates, clean_prices


def read_rows(zip_path, member):
    with ZipFile(zip_path) as zf:
        workbook = load_workbook(BytesIO(zf.read(member)), read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    headers = list(rows[0])
    return [dict(zip(headers, row)) for row in rows[1:] if any(cell is not None for cell in row)]


def read_wide_table(zip_path, member):
    with ZipFile(zip_path) as zf:
        workbook = load_workbook(BytesIO(zf.read(member)), read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    codes = list(rows[0])[1:]
    names = list(rows[1])[1:]
    dates = []
    series = {code: {"name": name, "prices": []} for code, name in zip(codes, names)}
    for row in rows[2:]:
        if row[0] in (None, ""):
            continue
        dates.append(excel_to_datetime(row[0]))
        for code, value in zip(codes, row[1:]):
            series[code]["prices"].append(to_float(value))
    return {"dates": dates, "series": series}


def choose_stock_universe(stock_info_rows, stock_table, n=5):
    info_by_code = {row["代码"]: row for row in stock_info_rows}
    ranked = []
    full_history = len(stock_table["dates"])
    for code, item in stock_table["series"].items():
        info = info_by_code.get(code, {})
        weight = to_float(info.get("权重", 0.0))
        history = sum(1 for price in item["prices"] if to_float(price) > 0)
        industry = info.get("Wind一级行业") or info.get("申银万国一级行业") or "未知"
        ranked.append((history == full_history, history, weight, industry, code, item["name"]))
    ranked.sort(reverse=True)
    chosen = []
    seen_industries = set()
    for is_full, history, weight, industry, code, name in ranked:
        if industry in seen_industries:
            continue
        chosen.append({"code": code, "name": name, "industry": industry, "weight": weight, "history": history})
        seen_industries.add(industry)
        if len(chosen) == n:
            return chosen
    for is_full, history, weight, industry, code, name in ranked:
        if any(row["code"] == code for row in chosen):
            continue
        chosen.append({"code": code, "name": name, "industry": industry, "weight": weight, "history": history})
        if len(chosen) == n:
            break
    return chosen
