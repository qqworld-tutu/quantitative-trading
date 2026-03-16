from datetime import datetime, timedelta
from io import BytesIO
from zipfile import ZipFile

from openpyxl.reader.excel import load_workbook

from assignment2.config import BENCHMARK, DATA_ZIP, STOCKS


def to_float(value):
    if value in (None, ""):
        return 0.0
    return float(value)


def excel_to_datetime(value):
    if isinstance(value, datetime):
        return value
    if isinstance(value, (int, float)):
        return datetime(1899, 12, 30) + timedelta(days=float(value))
    if isinstance(value, str):
        for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d"):
            try:
                return datetime.strptime(value, fmt)
            except ValueError:
                pass
    raise ValueError(f"unsupported datetime value: {value!r}")


def read_rows(zip_path, member):
    with ZipFile(zip_path) as zf:
        workbook = load_workbook(BytesIO(zf.read(member)), read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    headers = list(rows[0])
    return [dict(zip(headers, row)) for row in rows[1:] if any(cell is not None for cell in row)]


def read_wide_table(zip_path, member):
    with ZipFile(zip_path) as zf:
        workbook = load_workbook(BytesIO(zf.read(member)), read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    codes = list(rows[0])[1:]
    names = list(rows[1])[1:]
    dates = []
    series = {code: {"name": name, "prices": []} for code, name in zip(codes, names)}
    for row in rows[2:]:
        if row[0] in (None, ""):
            continue
        dates.append(excel_to_datetime(row[0]))
        for code, value in zip(codes, row[1:]):
            series[code]["prices"].append(to_float(value))
    return {"dates": dates, "series": series}


def align_price_table(table):
    common_dates = sorted(set.intersection(*(set(series) for series in table.values())))
    matrix = {name: [table[name][dt] for dt in common_dates] for name in table}
    return common_dates, matrix


def split_by_ratio(items, train_ratio, valid_ratio):
    n = len(items)
    i = int(n * train_ratio)
    j = i + int(n * valid_ratio)
    return items[:i], items[i:j], items[j:]


def _stock_prices():
    wide = read_wide_table(DATA_ZIP, "股票类/沪深300成分股收盘价日数据.xlsx")
    dates = wide["dates"]
    table = {}
    for code, name in STOCKS:
        prices = wide["series"][code]["prices"]
        table[name] = {dt: to_float(price) for dt, price in zip(dates, prices) if to_float(price) > 0}
    return table


def _benchmark_prices():
    code, name = BENCHMARK
    rows = read_rows(DATA_ZIP, "市场指数类/沪深300指数日数据.xlsx")
    return {excel_to_datetime(row["日期"]): to_float(row["收盘价(元)"]) for row in rows if to_float(row["收盘价(元)"]) > 0}


def load_aligned_prices():
    stock_table = _stock_prices()
    dates, prices = align_price_table(stock_table)
    benchmark_map = _benchmark_prices()
    benchmark = [benchmark_map[dt] for dt in dates]
    return dates, prices, benchmark


def split_dataset(dates, prices, benchmark, train_ratio=0.5, valid_ratio=0.1):
    train_dates, valid_dates, test_dates = split_by_ratio(dates, train_ratio, valid_ratio)
    i = len(train_dates)
    j = i + len(valid_dates)
    split_prices = {
        "train": {name: values[:i] for name, values in prices.items()},
        "valid": {name: values[i:j] for name, values in prices.items()},
        "test": {name: values[j:] for name, values in prices.items()},
    }
    benchmark_parts = {
        "train": benchmark[:i],
        "valid": benchmark[i:j],
        "test": benchmark[j:],
    }
    return {
        "dates": {"train": train_dates, "valid": valid_dates, "test": test_dates},
        "prices": split_prices,
        "benchmark": benchmark_parts,
    }
