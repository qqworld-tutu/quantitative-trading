from datetime import datetime, timedelta
from io import BytesIO
from zipfile import ZipFile

import numpy as np
from openpyxl.reader.excel import load_workbook

from assignment3.config import DAILY_RF, DATA_ZIP, STOCKS


def to_float(value):
    if value in (None, ""):
        return 0.0
    return float(value)


def excel_to_datetime(value):
    if isinstance(value, datetime):
        return value
    if isinstance(value, (int, float)):
        return datetime(1899, 12, 30) + timedelta(days=float(value))
    if isinstance(value, str):
        for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d"):
            try:
                return datetime.strptime(value, fmt)
            except ValueError:
                pass
    raise ValueError(f"unsupported datetime value: {value!r}")


def read_rows(zip_path, member):
    with ZipFile(zip_path) as zf:
        workbook = load_workbook(BytesIO(zf.read(member)), read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    headers = list(rows[0])
    return [dict(zip(headers, row)) for row in rows[1:] if any(cell is not None for cell in row)]


def read_wide_table(zip_path, member):
    with ZipFile(zip_path) as zf:
        workbook = load_workbook(BytesIO(zf.read(member)), read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    codes = list(rows[0])[1:]
    names = list(rows[1])[1:]
    dates = []
    series = {code: {"name": name, "prices": []} for code, name in zip(codes, names)}
    for row in rows[2:]:
        if row[0] in (None, ""):
            continue
        dates.append(excel_to_datetime(row[0]))
        for code, value in zip(codes, row[1:]):
            series[code]["prices"].append(to_float(value))
    return {"dates": dates, "series": series}


def split_common_dates(table):
    common_dates = sorted(set.intersection(*(set(series.keys()) for series in table.values())))
    values = {name: [series[dt] for dt in common_dates] for name, series in table.items()}
    return common_dates, values


def find_month_ends(dates):
    result = []
    for idx, dt in enumerate(dates):
        if idx == len(dates) - 1 or dates[idx + 1].month != dt.month:
            result.append(idx)
    return result


def log_returns(prices):
    arr = np.asarray(prices, dtype=float)
    return np.diff(np.log(arr))


def load_prices_and_returns():
    stock_wide = read_wide_table(DATA_ZIP, "股票类/沪深300成分股收盘价日数据.xlsx")
    stock_maps = {}
    for code, name in STOCKS:
        stock_maps[name] = {
            dt: to_float(price)
            for dt, price in zip(stock_wide["dates"], stock_wide["series"][code]["prices"])
            if to_float(price) > 0
        }
    index_rows = read_rows(DATA_ZIP, "市场指数类/沪深300指数日数据.xlsx")
    stock_dates, stock_prices = split_common_dates(stock_maps)
    index_map = {
        excel_to_datetime(row["日期"]): to_float(row["收盘价(元)"])
        for row in index_rows
        if to_float(row["收盘价(元)"]) > 0
    }
    common_dates = [dt for dt in stock_dates if dt in index_map]
    prices = {name: [stock_maps[name][dt] for dt in common_dates] for name in stock_maps}
    benchmark_prices = [index_map[dt] for dt in common_dates]
    asset_returns = np.column_stack([log_returns(prices[name]) for name in prices])
    benchmark_returns = log_returns(benchmark_prices)
    return {
        "dates": common_dates[1:],
        "asset_names": list(prices),
        "prices": prices,
        "asset_returns": asset_returns,
        "benchmark_returns": benchmark_returns,
        "risk_free": np.full(len(common_dates) - 1, DAILY_RF),
    }

